# Shows the top tracks for a user

import spotipy
import json
import openpyxl
from spotipy.oauth2 import SpotifyOAuth

scope = 'user-top-read'
sp = spotipy.Spotify(auth_manager=SpotifyOAuth(scope=scope))


user = "longboq"

playlists = sp.user_playlists(user)


# loop through playlists
for pl in playlists['items']:
    pl_name = pl['name']
    pl_url = pl['external_urls']['spotify']
    pl_uri = pl['uri']

    wb = openpyxl.Workbook()
    ws = wb.active
    cidx = 1
    for item in ["Title", "Album", "Artist", "Danceability", "Energy", "Key", "Loudness", "Mode", "Speechiness", "Acousticness", "Instrumentalness", "Liveness", "Valence", "Tempo"]:
        ws.cell(row=1, column=cidx).value = item
        cidx = cidx+1
    wb.save(f"dat/{pl_name}.xlsx")
    wb.close()
        
    ofs = 0
    while True:
        tracks = sp.playlist_items(pl_uri, offset=ofs, fields='items.track.id,total', additional_types=['track'])
        # check if playlist done
        if len(tracks['items']) == 0:
            break

        ofs = ofs + len(tracks['items'])
        # looping cond complete

        # loop through tracks
        for track in tracks['items']:
            tid = f"spotify:track:{track['track']['id']}"
            try:
                tr = sp.track(tid)
            except:
                print(f"track skipped: [{tr['name']}]")
                continue
            
            # extract track info
            name_tr = tr['name']
            name_alb = tr['album']['name']
            name_art = tr['artists'][0]['name']

            features = sp.audio_features(tid)
            fo = features[0]

            crit_feat = [fo['danceability'], fo['energy'], fo['key'], fo['loudness'], fo['mode'], fo['speechiness'], fo['acousticness'], fo['instrumentalness'], fo['liveness'], fo['valence'], fo['tempo']]

            payload = [name_tr, name_alb, name_art]
            payload = payload + crit_feat

            # now print into spreadsheet
            wb = openpyxl.load_workbook(f"dat/{pl_name}.xlsx")
            ws = wb.active

            # find first unused row 
            r = 1
            while (ws.cell(row=r, column=1).value != None):
                r = r + 1

            # print data in col
            for c in range(1,15):
                ws.cell(row=r, column=c, value=payload[c-1])

            wb.save(f"dat/{pl_name}.xlsx")
            wb.close()

    #break


print("Exited normally")
